"""
Phiên bản tối ưu tốc độ – v7
============================
Kế thừa toàn bộ WIN 1-13 của v6, bổ sung:

  [WIN 14] Eviction pass (post-processing gom YB về đúng block)
            Vấn đề: WC rải rác nhiều bay → drain-first vẫn buộc phải mở
            nhiều YB do không có đủ hàng cùng WC trong 1-2 bay.
            Giải pháp: Sau khi greedy loop xong, nếu thực tế dùng > MAX_YBS
            thì tiến hành eviction:
              - Giữ MAX YBs nhiều cont nhất (bay gần, hàng đông).
              - Un-pick container ở YB bị evict: trả về avail của block.
              - Re-block các container bên dưới bị ảnh hưởng.
              - Rollback opened_ybs cho YB evicted.
              - Tăng remaining → caller tự phân bổ số cont thiếu sang block khác.
            Kết quả: A02 chỉ dùng tối đa MAX YBs, mấy cont lẻ tẻ chuyển sang A03/B01.

  Hằng số điều chỉnh:
    MAX_YBS_PER_ASSIGN = 3  (TUNING FLAGS – thử 1, 2, hoặc 4)
"""
import io
import os
import sys
import subprocess
import tempfile
import re
import pandas as pd
import pulp
from collections import defaultdict

try:
    import streamlit as st
    _HAS_ST = True
except ImportError:
    _HAS_ST = False

def _log(msg: str):
    """In ra terminal VÀ hiện lên Streamlit UI (nếu có)."""
    print(msg)
    if _HAS_ST:
        st.write(msg)

# ============================================================
# TUNING FLAGS
# ============================================================
FAST_MODE         = True   # True = bỏ block_bay_wc khỏi MIP (khuyến nghị)
SOLVER_TIME_LIMIT = 60    # giây – 0 clashes thường tìm thấy trước 40s
MAX_YBS_PER_ASSIGN    = 2 # Tối đa YB mở cho 1 lần gán (2=chặt, 3=vừa, 4=lỏng)
ALT_BLOCK_ON_EVICTION = True
MAX_CONT_PER_BLOCK_BAY = 50
                          # Số container tối đa 1 block được phép cấp cho
                          # 1 vessel bay (tính tổng mọi giờ + mọi WC).
                          # Ràng buộc được áp cả vào MIP lẫn pick_n.
                          # Đặt None hoặc 0 để tắt giới hạn này.
                          # True  = khi block gốc bị evict, deferred thử block
                          #         khác cùng ST/POD trước → cont lẻ sang A03/B01
                          # False = hành vi cũ: luôn retry cùng block gốc



# ============================================================
# SOLVER HELPER
# ============================================================
def _n_threads():
    try:    return max(1, os.cpu_count() or 1)
    except: return 1
# ============================================================
# Sử dụng HiGHS Python API (an toàn, không cần file .exe)
# ============================================================
def _make_higsh_solver(time_limit):
    """HiGHS API (không cần file exe)."""
    import pulp
    try:
        solver = pulp.HiGHS(
            msg=True,
            timeLimit=time_limit,
            options={"parallel": "on", "threads": str(_n_threads())}
        )
        # test
        tp = pulp.LpProblem("_t", pulp.LpMinimize)
        tv = pulp.LpVariable("_v"); tp += tv; tp += tv >= 0
        tp.solve(solver)
        _log(f"[Solver] ✅ HiGHS (API) – {time_limit}s limit")
        return solver
    except Exception as e:
        _log(f"[Solver] ❌ HiGHS API: {e}")
        raise RuntimeError("HiGHS API unavailable")

# ============================================================
# MOVE HOUR SORT KEY
# ============================================================
_DAY_RANK = {'MO': 0, 'TU': 1, 'WE': 2, 'TH': 3, 'FR': 4, 'SA': 5, 'SU': 6}

def _hour_sort_key(h: str):
    s = str(h).strip().lstrip('+')
    if len(s) >= 2:
        dr = _DAY_RANK.get(s[:2].upper(), 99)
        try:    ti = int(s[2:]) if s[2:] else 0
        except: ti = 0
        return (dr, ti)
    return (99, s)

# ============================================================
# PUBLIC API
# ============================================================
def run_optimization(file_input):

    # =========================================================
    # 1. ĐỌC DỮ LIỆU
    # =========================================================
    xls = pd.ExcelFile(file_input)

    df1 = pd.read_excel(xls, sheet_name='MOVEHOUR-WEIGHTCLASS', header=None)
    has_st_pod    = (str(df1.iloc[1, 2]).strip().upper() == 'ST')
    data_col_start = 4 if has_st_pod else 2

    sts_bay_map = {}
    for col in range(data_col_start, df1.shape[1]):
        sts = df1.iloc[0, col]; bay = df1.iloc[1, col]
        if pd.notna(sts) and pd.notna(bay):
            sts_bay_map[col] = (str(sts).strip(), str(bay).strip())

    demands = {}
    cur_hour = None
    for idx in range(2, df1.shape[0]):
        row = df1.iloc[idx]
        hour = row[0]
        if pd.isna(hour): hour = cur_hour
        else:             cur_hour = hour
        weight = row[1]
        if pd.isna(weight): continue
        weight = int(float(str(weight)))
        st_v  = str(row[2]).strip() if has_st_pod and pd.notna(row[2]) else ''
        pod_v = str(row[3]).strip() if has_st_pod and pd.notna(row[3]) else ''
        for col in range(data_col_start, df1.shape[1]):
            qty = row[col]
            if pd.notna(qty) and qty != '':
                qty = int(float(str(qty)))
                if qty > 0:
                    sts, bay = sts_bay_map[col]
                    key  = (hour, sts, bay)
                    dkey = (weight, st_v, pod_v)
                    demands.setdefault(key, {})
                    demands[key][dkey] = demands[key].get(dkey, 0) + qty

    _log(f"Demand format: {'WC+ST+POD' if has_st_pod else 'WC only'}")
    job_keys          = list(demands.keys())
    all_hours_sorted  = sorted({h for h, s, b in job_keys}, key=_hour_sort_key)
    hour_rank         = {h: i for i, h in enumerate(all_hours_sorted)}
    jobs_by_hour      = defaultdict(list)
    for h, s, b in job_keys: jobs_by_hour[h].append((s, b))
    jobs_by_bay       = defaultdict(list)
    for h, s, bay in job_keys: jobs_by_bay[bay].append((h, s, bay))

    df2 = pd.read_excel(xls, sheet_name='BLOCK-WEIGHT CLASS', header=0)
    col_names         = [str(c).strip() for c in df2.columns]
    has_st_pod_supply = (col_names[1].upper() == 'ST' and col_names[2].upper() == 'POD')
    wc_col_start      = 3 if has_st_pod_supply else 1

    supply = {}; blocks_set = set()
    for _, row in df2.iterrows():
        block = str(row.iloc[0]).strip()
        if block in ('nan', 'GRAND TOTAL', '') or not block: continue
        st_v  = str(row.iloc[1]).strip() if has_st_pod_supply else ''
        pod_v = str(row.iloc[2]).strip() if has_st_pod_supply else ''
        skey  = (block, st_v, pod_v)
        wc_dict = {}
        for wi, w in enumerate([1, 2, 3, 4, 5]):
            ci  = wc_col_start + wi
            val = row.iloc[ci] if ci < len(row) else None
            wc_dict[w] = int(val) if pd.notna(val) and val != '' else 0
        supply[skey] = wc_dict
        blocks_set.add(block)

    weight_classes = [1, 2, 3, 4, 5]
    blocks         = sorted(blocks_set)
    supply_keys    = [k for k in supply if any(supply[k][w] > 0 for w in weight_classes)]
    _log(f"Supply format: {'BLOCK+ST+POD' if has_st_pod_supply else 'BLOCK only'}")

    # ── Map (MOVE HOUR, BAY số) → VESSEL BAY ─────────────────────────
    # Cột J (index 9)  = VESSEL BAY, ví dụ "34A", "34B", "02A"
    # Cột O (index 14) = MOVE HOUR,  ví dụ "+SA0200"
    # Cột P (index 15) = BAY,        ví dụ "34", "02"  (text số thuần)
    #
    # Matching rule: BAY cột P ("34") khớp với phần số của VESSEL BAY cột J
    # ("34A" → "34", "34B" → "34") → lấy đúng vessel bay theo từng bay.
    # Key lookup: (move_hour_str, bay_numeric_str) → vessel_bay_str

    def _bay_num(val: str) -> str:
        """Trích phần số liên tiếp đầu tiên trong chuỗi bay.
        'BAY 34' → '34', '34A' → '34', '02B' → '02', '34' → '34'."""
        import re
        m = re.search(r'\d+', str(val).strip())
        return m.group() if m else ''

    move_hour_vessel_bay: dict = {}   # key: (mh_str, bay_num_str) → vb_str
    try:
        df_raw = pd.read_excel(xls, sheet_name='DATA', header=None)
        COL_VESSEL_BAY = 9   # J
        COL_MOVE_HOUR  = 14  # O
        COL_BAY        = 15  # P
        if df_raw.shape[1] > max(COL_VESSEL_BAY, COL_MOVE_HOUR, COL_BAY):
            for _, row in df_raw.iterrows():
                mh  = row.iloc[COL_MOVE_HOUR]
                vb  = row.iloc[COL_VESSEL_BAY]
                bay = row.iloc[COL_BAY]
                if pd.notna(mh) and pd.notna(vb) and pd.notna(bay):
                    mh_str  = str(mh).strip()
                    vb_str  = str(vb).strip()
                    bay_num = _bay_num(str(bay))
                    if mh_str and vb_str and bay_num:
                        move_hour_vessel_bay[(mh_str, bay_num)] = vb_str
        _log(f"VESSEL BAY map: {len(move_hour_vessel_bay)} entries.")
    except Exception as e:
        _log(f"Không build được VESSEL BAY map: {e}")

    # Sheet DATA
    container_data_available = False
    df_containers            = None
    try:
        df_containers = pd.read_excel(xls, sheet_name='DATA', header=0)
        cols = list(df_containers.columns)
        def find_col(cands):
            for c in cands:
                if c in cols: return c
            return None
        wc_src  = find_col(['YC', 'Unnamed: 1'])
        yp_src  = find_col(['YP', 'Unnamed: 2'])
        id_src  = find_col(['ID', 'Unnamed: 3'])
        st_src  = find_col(['ST'])
        pod_src = find_col(['POD'])
        if wc_src and yp_src and all(c in cols for c in ['YB', 'YR', 'YT']):
            df_containers = df_containers.dropna(
                subset=[wc_src, yp_src, 'YB', 'YR', 'YT']).copy()
            df_containers['REAL_WC']      = df_containers[wc_src].astype(float).astype(int)
            df_containers['YARD_POS']     = df_containers[yp_src].astype(str).str.strip()
            df_containers['REAL_CONT_ID'] = (
                df_containers[id_src].fillna('').astype(str).str.strip() if id_src else '')
            df_containers['CONT_ST']  = (
                df_containers[st_src].fillna('').astype(str).str.strip() if st_src else '')
            df_containers['CONT_POD'] = (
                df_containers[pod_src].fillna('').astype(str).str.strip() if pod_src else '')
            df_containers['YARD'] = df_containers['YARD'].astype(str).str.strip()
            for c in ['YB', 'YR', 'YT']:
                df_containers[c] = df_containers[c].astype(float).astype(int)
            container_data_available = True
            _log(f"Container DATA: {len(df_containers)} rows.")
    except Exception as e:
        _log(f"No DATA sheet. ({e})")

    # Stacking
    yb_wc_supply = {}; stack_ordering = {}; blocking_pairs = []
    if container_data_available:
        df_c = df_containers[['YARD', 'YB', 'YR', 'YT', 'REAL_WC',
                               'YARD_POS', 'REAL_CONT_ID',
                               'CONT_ST', 'CONT_POD']].copy()
        for block in blocks:
            bdf = df_c[df_c['YARD'] == block]
            if bdf.empty: continue
            yb_wc_supply[block] = {}; stack_ordering[block] = {}
            for yb, yb_df in bdf.groupby('YB'):
                yb_wc_supply[block][yb] = {
                    wc: int(cnt) for wc, cnt in yb_df.groupby('REAL_WC').size().items()}
                stack_ordering[block][yb] = {}
                for yr, yr_df in yb_df.groupby('YR'):
                    ordered = yr_df.sort_values('YT', ascending=False)[
                        ['YT', 'REAL_WC']].values.tolist()
                    stack_ordering[block][yb][yr] = [(int(t), int(w)) for t, w in ordered]
                for yr, tiers in stack_ordering[block][yb].items():
                    wcs_above = []
                    for tier, wc in tiers:
                        for prev_wc, prev_tier in wcs_above:
                            if prev_wc != wc:
                                blocking_pairs.append(
                                    (block, yb, yr, prev_tier, prev_wc, tier, wc))
                        wcs_above.append((wc, tier))
        _log(f"Stacking: {len(blocking_pairs)} blocking pairs.")

    # =========================================================
    # 2. DEMAND / SUPPLY CHECK
    # =========================================================
    total_demand = defaultdict(int)
    for job in job_keys:
        for dkey, qty in demands[job].items():
            total_demand[dkey] += qty
    total_supply = defaultdict(int)
    for skey in supply_keys:
        b, st_v, pod_v = skey
        for w in weight_classes:
            total_supply[(w, st_v, pod_v)] += supply[skey][w]

    ok = True
    for k in set(list(total_demand) + list(total_supply)):
        d = total_demand.get(k, 0); s = total_supply.get(k, 0)
        if d != s:
            _log(f"  ERROR Mismatch WC={k[0]} ST={k[1]} POD={k[2]}: "
                  f"demand={d}, supply={s}")
            ok = False
    if not ok:
        raise ValueError("Demand/supply mismatch.")
    _log("Demand/supply balanced OK.")

    # =========================================================
    # 3. XÂY DỰNG MIP
    # =========================================================
    _log(f"[MIP] FAST_MODE={'ON' if FAST_MODE else 'OFF'}")

    # Pre-index supply blocks by (st, pod)
    supply_blocks_by_st_pod = defaultdict(list)
    for b, st_v, pod_v in supply_keys:
        supply_blocks_by_st_pod[(st_v, pod_v)].append(b)

    # ── [WIN 1] Prune y_vars: chỉ tạo (job, block) có supply khớp ──
    feasible_blocks_for_job = {}
    n_fixed = 0
    for (h, s, bay) in job_keys:
        fblocks = set()
        for dkey in demands[(h, s, bay)]:
            _, st_v, pod_v = dkey
            fblocks.update(supply_blocks_by_st_pod.get((st_v, pod_v), []))
        feasible_blocks_for_job[(h, s, bay)] = sorted(fblocks)

    y_vars = {}
    for (h, s, bay) in job_keys:
        fblocks = feasible_blocks_for_job[(h, s, bay)]
        for b in fblocks:
            y_vars[(h, s, bay, b)] = pulp.LpVariable(
                f"y_{h}_{s}_{bay}_{b}", cat='Binary')
        # [WIN 3] Fix trivial assignments (chỉ 1 block khả thi)
        if len(fblocks) == 1:
            y_vars[(h, s, bay, fblocks[0])].lowBound = 1
            n_fixed += 1

    if n_fixed:
        _log(f"[MIP] Fixed {n_fixed} jobs (single feasible block).")

    # x_vars – chỉ tạo khi y_var tương ứng tồn tại
    x_vars = {}
    for (h, s, bay) in job_keys:
        for dkey in demands[(h, s, bay)]:
            w, st_v, pod_v = dkey
            for b in supply_blocks_by_st_pod.get((st_v, pod_v), []):
                if (h, s, bay, b) in y_vars:
                    x_vars[(h, s, bay, b, dkey)] = pulp.LpVariable(
                        f"x_{h}_{s}_{bay}_{b}_{w}_{st_v}_{pod_v}",
                        lowBound=0, cat='Integer')

    prob = pulp.LpProblem("Minimize_Clashes", pulp.LpMinimize)

    CLASH_W = 70.0; SINGLE_W = 50.0; SPREAD_W = 50.0; BAY_SINGLE_W = 50.0

    # ── [WIN 2+4] Xóa u_vars, viết e trực tiếp + tight upper bound ──
    e_vars = {}
    for h in jobs_by_hour:
        for b in blocks:
            y_list = [y_vars[(h, s, bay, b)]
                      for (s, bay) in jobs_by_hour[h]
                      if (h, s, bay, b) in y_vars]
            if not y_list:
                continue   # [WIN 5] bỏ qua cặp (h,b) không có y_vars
            e = pulp.LpVariable(f"e_{h}_{b}", lowBound=0, cat='Integer')
            e_vars[(h, b)] = e
            ysum = pulp.lpSum(y_list)
            prob += e >= ysum - 1                 # lower bound
            prob += e <= len(y_list) - 1          # [WIN 4] tight upper bound

    # single_block per job
    single_block = {}
    for (h, s, bay) in job_keys:
        fblocks = feasible_blocks_for_job[(h, s, bay)]
        sb = pulp.LpVariable(f"sb_{h}_{s}_{bay}",
                             lowBound=0, upBound=1, cat='Continuous')
        single_block[(h, s, bay)] = sb
        prob += sb >= 2 - pulp.lpSum(y_vars[(h, s, bay, b)] for b in fblocks)

    # ── [WIN 5] Prune block_bay: chỉ tạo cho cặp (b, bay) có y_vars ──
    all_bays     = sorted({bay for (_, _, bay) in job_keys})
    relevant_bb  = set()
    for (h, s, bay) in job_keys:
        for b in feasible_blocks_for_job[(h, s, bay)]:
            relevant_bb.add((b, bay))

    block_bay = {}
    for (b, bay) in sorted(relevant_bb):
        var = pulp.LpVariable(f"bb_{b}_{bay}", cat='Binary')
        block_bay[(b, bay)] = var
        for (h, s, bj) in jobs_by_bay[bay]:
            if (h, s, bay, b) in y_vars:
                prob += var >= y_vars[(h, s, bay, b)]

    blocks_per_bay = defaultdict(list)
    for (b, bay) in relevant_bb:
        blocks_per_bay[bay].append(b)

    bay_single = {}
    for bay in all_bays:
        blist = blocks_per_bay.get(bay, [])
        if not blist:
            continue
        var = pulp.LpVariable(f"bs_{bay}", lowBound=0, upBound=1, cat='Continuous')
        bay_single[bay] = var
        bb_sum = pulp.lpSum(block_bay.get((b, bay), 0) for b in blist)
        prob += var >= 2 - bb_sum
        prob += bb_sum >= 2

    # block_bay_wc – chỉ khi FAST_MODE=False
    block_bay_wc = {}
    if not FAST_MODE:
        x_by_bbw = defaultdict(list)
        for (h, s, bay, b, dkey), xvar in x_vars.items():
            x_by_bbw[(b, bay, dkey[0])].append((xvar, demands[(h, s, bay)][dkey]))
        for b in blocks:
            for bay in all_bays:
                for wc in weight_classes:
                    entries = x_by_bbw.get((b, bay, wc), [])
                    if not entries: continue
                    var = pulp.LpVariable(f"bbw_{b}_{bay}_{wc}", cat='Binary')
                    block_bay_wc[(b, bay, wc)] = var
                    for xvar, d in entries:
                        prob += var >= xvar / (d + 0.1)

    # Objective
    obj = (CLASH_W      * pulp.lpSum(e_vars.values())      +
           SINGLE_W     * pulp.lpSum(single_block.values()) +
           SPREAD_W     * pulp.lpSum(block_bay.values())    +
           BAY_SINGLE_W * pulp.lpSum(bay_single.values()))
    if block_bay_wc:
        obj += 2.0 * pulp.lpSum(block_bay_wc.values())
    prob += obj

    # Demand constraints
    for (h, s, bay) in job_keys:
        for dkey, d in demands[(h, s, bay)].items():
            w, st_v, pod_v = dkey
            x_list = [x_vars[(h, s, bay, b, dkey)]
                      for b in supply_blocks_by_st_pod.get((st_v, pod_v), [])
                      if (h, s, bay, b, dkey) in x_vars]
            if x_list:
                prob += pulp.lpSum(x_list) == d

    # Supply constraints
    x_by_supply = defaultdict(list)
    for (h, s, bay, b, dkey), xvar in x_vars.items():
        w, st_v, pod_v = dkey
        x_by_supply[(b, st_v, pod_v, w)].append(xvar)
    for skey in supply_keys:
        b, st_v, pod_v = skey
        for w in weight_classes:
            xl = x_by_supply.get((b, st_v, pod_v, w), [])
            if xl:
                prob += pulp.lpSum(xl) <= supply[skey][w]

    # Linking x <= d * y
    for (h, s, bay) in job_keys:
        for dkey, d in demands[(h, s, bay)].items():
            for b in supply_blocks_by_st_pod.get((dkey[1], dkey[2]), []):
                key = (h, s, bay, b, dkey)
                if key in x_vars:
                    prob += x_vars[key] <= d * y_vars[(h, s, bay, b)]

    # ── MAX_CONT_PER_BLOCK_BAY constraint ──────────────────────────────
    # Tổng container block b cấp cho vessel bay (mọi giờ, mọi WC) <= giới hạn
    if MAX_CONT_PER_BLOCK_BAY:
        x_by_block_bay = defaultdict(list)
        for (h, s, bay, b, dkey), xvar in x_vars.items():
            x_by_block_bay[(b, bay)].append(xvar)
        for (b, bay), xlist in x_by_block_bay.items():
            prob += (pulp.lpSum(xlist) <= MAX_CONT_PER_BLOCK_BAY,
                     f"cap_bb_{b}_{bay}")

    t_build = len(prob.variables()); nc = len(prob.constraints)
    _log(f"Model: {t_build} biến, {nc} constraints")

 # =========================================================
    # GỌI SOLVER – Ưu tiên HiGHS API, fallback CBC subprocess
    # =========================================================
    try:
        solver = _make_higsh_solver(SOLVER_TIME_LIMIT)
        prob.solve(solver)
        status_str = pulp.LpStatus[prob.status]
    except RuntimeError as e:
        _log(f"[CBC] Fallback vì HiGHS không khả dụng: {e}")
        status_str = _run_cbc_with_timeout(prob, max(300, SOLVER_TIME_LIMIT * 2))
        prob.status = {
            "Optimal": pulp.LpStatusOptimal,
            "NotSolved": pulp.LpStatusNotSolved,
            "Infeasible": pulp.LpStatusInfeasible
        }.get(status_str, pulp.LpStatusNotSolved)

    # Kiểm tra có lời giải hay không
    feasible_found = False
    for yvar in y_vars.values():
        if pulp.value(yvar) is not None:
            feasible_found = True
            break
    if not feasible_found:
        raise RuntimeError(
            "Không tìm được lời giải khả thi trong thời gian cho phép.\n"
            "Hãy tăng SOLVER_TIME_LIMIT (hiện tại {} giây).".format(SOLVER_TIME_LIMIT)
        )
    # =========================================================
    # 4. KẾT QUẢ VÀ GÁN CONTAINER
    # =========================================================
    result_rows = []
    for (h, s, bay, b), yvar in y_vars.items():
        yv = pulp.value(yvar)
        if yv is not None and yv > 0.5:
            for dkey in demands[(h, s, bay)]:
                xkey = (h, s, bay, b, dkey)
                if xkey not in x_vars: continue
                qty = pulp.value(x_vars[xkey])
                if qty is not None and qty > 0.5:
                    w, st_v, pod_v = dkey
                    result_rows.append({
                        'MOVE HOUR': h, 'STS': s, 'BAY': bay,
                        'ASSIGNED BLOCK': b, 'WEIGHT CLASS': w,
                        'ST': st_v, 'POD': pod_v,
                        'QUANTITIES': int(round(qty))
                    })
    df_result = pd.DataFrame(result_rows)
    df_result['_sort_hr'] = df_result['MOVE HOUR'].map(hour_rank)
    df_result.sort_values(['_sort_hr', 'STS', 'BAY', 'ASSIGNED BLOCK'], inplace=True)
    df_result.drop(columns=['_sort_hr'], inplace=True)

    # Container assignment
    df_result_detail = []
    if container_data_available:
        # Build pool
        pool = defaultdict(list)
        for row in df_containers[['YARD', 'YB', 'YR', 'YT', 'REAL_WC',
                                   'YARD_POS', 'REAL_CONT_ID',
                                   'CONT_ST', 'CONT_POD']].itertuples(index=False):
            c = {'yb': int(row.YB), 'yr': int(row.YR), 'yt': int(row.YT),
                 'wc': int(row.REAL_WC), 'yard_pos': row.YARD_POS,
                 'real_cont_id': row.REAL_CONT_ID,
                 'st': row.CONT_ST, 'pod': row.CONT_POD,
                 'picked': False, 'pick_h': None}
            pool[row.YARD].append(c)

        # O(1) accessibility
        blocked_count = {}; below_map = {}
        for blk, conts in pool.items():
            stacks = defaultdict(list)
            for c in conts: stacks[(c['yb'], c['yr'])].append(c)
            for c in conts:
                above = [o for o in stacks[(c['yb'], c['yr'])]
                         if o is not c and o['yt'] > c['yt']]
                below = [o for o in stacks[(c['yb'], c['yr'])]
                         if o is not c and o['yt'] < c['yt']]
                blocked_count[id(c)] = len(above)
                below_map[id(c)]     = below

        avail = defaultdict(list)
        for blk, conts in pool.items():
            for c in conts:
                if blocked_count[id(c)] == 0:
                    avail[(blk, c['wc'], c['st'], c['pod'])].append(c)

        opened_ybs = set()
        # Tracker: số cont đã pick theo (block, vessel_bay) – enforce MAX_CONT_PER_BLOCK_BAY
        block_bay_picked: dict = defaultdict(int)

        def pick_n(block, wc, st_match, pod_match, qty,
                   h, s_job, bay_job, h_rank_val, result_list):
            """
            Cải tiến v7 – Eviction pass (post-processing gom YB):
              Nguyên nhân gốc: WC khác nhau buộc phải mở nhiều YB,
              drain-first không thể cản được nếu WC rải đều nhiều bay.

              Giải pháp v7:
              1. Chạy greedy pick như v6 (drain-first + row-complete).
              2. SAU KHI loop xong, kiểm tra: nếu thực sự dùng > MAX_YBS_PER_ASSIGN
                 yard-bay thì tiến hành EVICTION:
                 - Giữ lại MAX_YBS_PER_ASSIGN YBs có nhiều cont nhất (đứng gần nhau).
                 - Các YB còn lại: UN-PICK container, trả về avail của block này.
                 - Increment remaining để caller biết còn thiếu → chuyển sang block khác.
              3. Caller (vòng deferred) sẽ tự động phân bổ số cont thiếu sang block khác.

              Ưu điểm: block A02 chắc chắn chỉ dùng tối đa MAX YBs,
              các cont lẻ tẻ được trả về và phân bổ sang A03/B01.
            """
            remaining = qty
            akey = (block, wc, st_match, pod_match)
            av   = avail[akey]

            # Tính quota còn lại cho (block, vessel_bay) này
            if MAX_CONT_PER_BLOCK_BAY:
                already = block_bay_picked[(block, bay_job)]
                quota   = MAX_CONT_PER_BLOCK_BAY - already
                if quota <= 0:
                    # Block này đã đủ quota cho vessel bay → trả toàn bộ về deferred
                    return remaining
                remaining = min(remaining, quota)

            # local_opened: YBs mở trong lần này; newly_opened: YBs do lần này mở mới
            local_opened:  set  = set()
            newly_opened:  set  = set()   # để rollback opened_ybs khi evict
            local_picked:  dict = defaultdict(int)
            my_picks:      list = []      # (container_obj, entry_dict) — ghi cuối cùng

            while remaining > 0:
                if not av:
                    break

                yb_cnt  = defaultdict(int)
                row_cnt = defaultdict(int)
                for c in av:
                    yb_cnt[c['yb']] += 1
                    row_cnt[(c['yb'], c['yr'])] += 1

                has_local_avail = any(c['yb'] in local_opened for c in av)

                best_i    = None
                best_score = None

                for i, c in enumerate(av):
                    yb = c['yb']
                    yr = c['yr']

                    if yb not in local_opened:
                        if len(local_opened) >= MAX_YBS_PER_ASSIGN:
                            continue
                        if has_local_avail:
                            continue

                    s_drain        = -local_picked[yb]
                    s_local        = 0 if yb in local_opened else 1
                    s_open         = 0 if (block, yb) in opened_ybs else 1
                    s_row_complete = -1 if row_cnt[(yb, yr)] <= remaining else 0
                    s_yb_dense     = -yb_cnt[yb]
                    s_yb           = yb
                    s_yr           = yr
                    s_yt           = -c['yt']

                    score = (s_drain, s_local, s_open,
                             s_row_complete, s_yb_dense, s_yb, s_yr, s_yt)

                    if best_score is None or score < best_score:
                        best_score = score
                        best_i = i

                if best_i is None:
                    break

                best = av.pop(best_i)
                best['picked'] = True
                best['pick_h'] = h

                if (block, best['yb']) not in opened_ybs:
                    newly_opened.add(best['yb'])
                local_opened.add(best['yb'])
                local_picked[best['yb']] += 1
                opened_ybs.add((block, best['yb']))

                for lower in below_map.get(id(best), []):
                    if lower['picked']:
                        continue
                    blocked_count[id(lower)] -= 1
                    if blocked_count[id(lower)] == 0:
                        avail[(block, lower['wc'],
                               lower['st'], lower['pod'])].append(lower)

                my_picks.append((best, {
                    'MOVE HOUR': h, 'CONTAINER ID': best['real_cont_id'],
                    'ST': best['st'], 'POD': best['pod'],
                    'STS': s_job, 'BAY': bay_job,
                    'VESSEL BAY': move_hour_vessel_bay.get(
                        (str(h).strip(), _bay_num(str(bay_job))), ''),
                    'ASSIGNED BLOCK': block, 'WEIGHT CLASS': wc,
                    'QUANTITIES': qty,
                    'YB': best['yb'], 'YR': best['yr'], 'YT': best['yt'],
                    'YARD POSITION': best['yard_pos']
                }))
                remaining -= 1

            # ── Eviction pass ────────────────────────────────────────────
            # Đếm số cont đã lấy theo YB
            yb_pick_cnt: dict = defaultdict(int)
            for c_obj, _ in my_picks:
                yb_pick_cnt[c_obj['yb']] += 1

            if len(yb_pick_cnt) > MAX_YBS_PER_ASSIGN:
                # Sắp xếp: nhiều cont nhất → ít nhất, cùng số thì YB số nhỏ trước
                sorted_ybs = sorted(yb_pick_cnt,
                                    key=lambda yb: (-yb_pick_cnt[yb], yb))
                keep_ybs  = set(sorted_ybs[:MAX_YBS_PER_ASSIGN])
                evict_ybs = set(sorted_ybs[MAX_YBS_PER_ASSIGN:])

                kept = []
                for c_obj, entry in my_picks:
                    if c_obj['yb'] in keep_ybs:
                        kept.append((c_obj, entry))
                    else:
                        # ── Un-pick ──────────────────────────────────────
                        c_obj['picked'] = False
                        c_obj['pick_h'] = None
                        # Trả container về avail của block này
                        av.append(c_obj)
                        remaining += 1   # báo caller còn thiếu

                        # Re-block các container bên dưới đã bị unblock lúc pick
                        for lower in below_map.get(id(c_obj), []):
                            if lower['picked']:
                                continue
                            blocked_count[id(lower)] += 1
                            if blocked_count[id(lower)] > 0:
                                akey_l = (block, lower['wc'],
                                          lower['st'], lower['pod'])
                                try:
                                    avail[akey_l].remove(lower)
                                except ValueError:
                                    pass   # đã bị pop trước đó

                # Rollback opened_ybs cho các YB evicted mà ta vừa mở mới
                for yb in evict_ybs:
                    if yb in newly_opened:
                        opened_ybs.discard((block, yb))

                my_picks = kept
            # ─────────────────────────────────────────────────────────────

            # Chỉ ghi vào result_list sau khi eviction xong
            for _, entry in my_picks:
                result_list.append(entry)

            # Cập nhật tracker block_bay_picked
            if MAX_CONT_PER_BLOCK_BAY:
                block_bay_picked[(block, bay_job)] += len(my_picks)

            # Tính lại remaining thực tế (qty gốc - số đã pick thành công)
            actual_picked = len(my_picks)
            return qty - actual_picked

        # [WIN 7] sort once, iterate with hour filter
        df_rs = df_result.copy()
        df_rs['_hr'] = df_rs['MOVE HOUR'].map(hour_rank)
        df_rs.sort_values(['_hr', 'STS', 'BAY', 'ASSIGNED BLOCK', 'WEIGHT CLASS'],
                          inplace=True)
        # Group by hour for fast per-hour access
        rs_by_hour = defaultdict(list)
        for rec in df_rs.to_dict('records'):
            rs_by_hour[rec['MOVE HOUR']].append(rec)

        def _pick_with_alt(orig_b, wc, st_v, pod_v, qty,
                           h, s_job, bay_job, hrv):
            """
            Wrapper cho deferred retry khi ALT_BLOCK_ON_EVICTION=True:
              1. Thử tất cả block KHÁC cùng ST/POD (sorted để ổn định)
              2. Fallback block gốc nếu vẫn còn thiếu
            Trả về remaining sau khi đã thử hết các phương án.
            """
            if not ALT_BLOCK_ON_EVICTION:
                return pick_n(orig_b, wc, st_v, pod_v, qty,
                              h, s_job, bay_job, hrv, df_result_detail)

            alt_blocks = sorted(
                b2 for b2 in supply_blocks_by_st_pod.get((st_v, pod_v), [])
                if b2 != orig_b and b2 in pool
            )
            rem = qty
            for b2 in alt_blocks:
                if rem <= 0:
                    break
                rem = pick_n(b2, wc, st_v, pod_v, rem,
                             h, s_job, bay_job, hrv, df_result_detail)
            # Fallback: thử lại block gốc nếu chưa đủ
            if rem > 0:
                rem = pick_n(orig_b, wc, st_v, pod_v, rem,
                             h, s_job, bay_job, hrv, df_result_detail)
            return rem

        deferred = []
        for h in all_hours_sorted:
            hrv = hour_rank[h]
            for asg in rs_by_hour[h]:
                s, bay_job, b = asg['STS'], asg['BAY'], asg['ASSIGNED BLOCK']
                w     = int(asg['WEIGHT CLASS'])
                st_v  = str(asg.get('ST', '')).strip()
                pod_v = str(asg.get('POD', '')).strip()
                qty   = int(asg['QUANTITIES'])
                if b not in pool:
                    df_result_detail.append({
                        'MOVE HOUR': h, 'STS': s, 'BAY': bay_job,
                        'VESSEL BAY': move_hour_vessel_bay.get(
                            (str(h).strip(), _bay_num(str(bay_job))), ''),
                        'ASSIGNED BLOCK': b, 'WEIGHT CLASS': w,
                        'CONTAINER ID': '', 'ST': '', 'POD': '',
                        'QUANTITIES': qty,
                        'YB': '', 'YR': '', 'YT': '', 'YARD POSITION': ''
                    })
                    continue
                rem = pick_n(b, w, st_v, pod_v, qty,
                             h, s, bay_job, hrv, df_result_detail)
                if rem > 0:
                    deferred.append({'b': b, 'wc': w, 'st': st_v, 'pod': pod_v,
                                     'qty': rem, 'h_orig': h, 's': s,
                                     'bay': bay_job, 'h_rank_min': hrv})

            still_def = []
            for d in deferred:
                rem = _pick_with_alt(d['b'], d['wc'], d['st'], d['pod'], d['qty'],
                                     h, d['s'], d['bay'], hrv)
                if rem > 0:
                    d2 = d.copy(); d2['qty'] = rem; still_def.append(d2)
            deferred = still_def

        rh = sum(d['qty'] for d in deferred)
        if rh > 0:
            _log(f"  Re-handling: {rh} containers.")
            for d in deferred:
                _log(f"    Block {d['b']} WC{d['wc']} x{d['qty']} "
                      f"(from {d['h_orig']})")
        else:
            _log("  All containers assigned – no re-handling.")

        df_result_detail = pd.DataFrame(df_result_detail)
    else:
        df_result_detail = df_result.copy()
        df_result_detail.insert(1, 'CONTAINER ID', '')
        df_result_detail.insert(2, 'ST', '')
        df_result_detail.insert(3, 'POD', '')
        df_result_detail['YB'] = ''
        df_result_detail['YR'] = ''
        df_result_detail['YT'] = ''
        df_result_detail['YARD POSITION'] = ''

    df_result_detail['_sort_hr'] = df_result_detail['MOVE HOUR'].map(hour_rank)
    df_result_detail.sort_values(
        ['_sort_hr', 'STS', 'BAY', 'ASSIGNED BLOCK',
         'WEIGHT CLASS', 'YB', 'YR', 'YT'], inplace=True)
    df_result_detail.drop(columns=['_sort_hr'], inplace=True)

    # =========================================================
    # 5. CLASH
    # =========================================================
    clash_details = []
    total_clashes = 0
    for (h, b), evar in e_vars.items():
        e_val = pulp.value(evar)
        if e_val is not None and e_val > 0.5:
            total_clashes += e_val
            jobs = [f"{s}@{bay}" for (s, bay) in jobs_by_hour.get(h, [])
                    if (h, s, bay, b) in y_vars
                    and pulp.value(y_vars[(h, s, bay, b)]) is not None
                    and pulp.value(y_vars[(h, s, bay, b)]) > 0.5]
            u_val = len(jobs) + int(e_val)   # u = e + 1 conceptually
            clash_details.append({
                'MOVE HOUR': h, 'BLOCK': b,
                'SỐ LƯỢNG BAY (u)': u_val,
                'CLASH (e = u-1)': int(e_val),
                'DANH SÁCH JOB (STS@BAY)': ', '.join(jobs)
            })
    df_clash = pd.DataFrame(clash_details)
    if not df_clash.empty:
        df_clash['_sort_hr'] = df_clash['MOVE HOUR'].map(hour_rank)
        df_clash.sort_values(['_sort_hr', 'BLOCK'], inplace=True)
        df_clash.drop(columns=['_sort_hr'], inplace=True)
    _log(f"Total clashes: {total_clashes}")

    # =========================================================
    # 6. GHI EXCEL – openpyxl tối ưu
    # =========================================================
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

    def _side():
        return Side(border_style='thin', color='FF000000')
    def _border():
        s = _side()
        return Border(left=s, right=s, top=s, bottom=s)
    def _font(bold=False, color='FF000000', size=10):
        return Font(name='Calibri', bold=bold, color=color, size=size)
    def _fill(hex6):
        return PatternFill('solid', fgColor=hex6)
    def _align(h='center', v='center', wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    _OX = {
        'dark':  'FF1F4E79', 'mid':   'FF2E75B6',
        'light': 'FF9DC3E6', 'pale':  'FFD6E4F0',
        'alt':   'FFEBF3FB', 'white': 'FFFFFFFF',
    }
    BD = _border()

    _ST = {
        'hdr_dark':  (_font(bold=True, color='FFFFFFFF'), _fill(_OX['dark']),
                      _align(wrap=True), BD),
        'hdr_mid':   (_font(bold=True, color='FFFFFFFF'), _fill(_OX['mid']),
                      _align(wrap=True), BD),
        'hdr_light': (_font(bold=True),                   _fill(_OX['light']),
                      _align(wrap=True), BD),
        'hdr_pale':  (_font(bold=True),                   _fill(_OX['pale']),
                      _align(wrap=True), BD),
        'data_w':    (_font(), _fill(_OX['white']), _align(), BD),
        'data_a':    (_font(), _fill(_OX['alt']),   _align(), BD),
        'cont_list': (_font(size=9), _fill(_OX['pale']),
                      _align(h='left', v='top', wrap=True), BD),
    }

    def _apply(cell, key):
        f, fi, al, bo = _ST[key]
        cell.font = f; cell.fill = fi
        cell.alignment = al; cell.border = bo

    wb = openpyxl.Workbook(); wb.remove(wb.active)

    # CLASH sheet
    ws_clash   = wb.create_sheet('CLASH')
    hdrs_clash = ['MOVE HOUR', 'BLOCK', 'SỐ LƯỢNG BAY (u)',
                  'CLASH (e = u-1)', 'DANH SÁCH JOB (STS@BAY)']
    clash_col_w = [14, 12, 18, 18, 50]
    ws_clash.append(hdrs_clash)
    for ci, (hdr, cw) in enumerate(zip(hdrs_clash, clash_col_w), 1):
        cell = ws_clash.cell(row=1, column=ci)
        _apply(cell, 'hdr_dark')
        ws_clash.column_dimensions[get_column_letter(ci)].width = cw
    if not df_clash.empty:
        clash_recs = df_clash[hdrs_clash].to_dict('records')
        for ri, rec in enumerate(clash_recs, 2):
            ws_clash.append([rec.get(c, '') for c in hdrs_clash])
            sk = 'data_w' if ri % 2 == 0 else 'data_a'
            for ci in range(1, 6):
                _apply(ws_clash.cell(row=ri, column=ci), sk)
    else:
        ws_clash.append(['Không có clash nào xảy ra.'])
        ws_clash.merge_cells(start_row=2, start_column=1,
                             end_row=2,   end_column=5)
        _apply(ws_clash.cell(row=2, column=1), 'data_w')

    # Result column config
    core_cols     = ['MOVE HOUR', 'CONT LIST', 'CONTAINER ID', 'ST', 'POD',
                     'STS', 'VESSEL BAY', 'BAY', 'ASSIGNED BLOCK',
                     'WEIGHT CLASS', 'QUANTITIES']
    position_cols = ['YB', 'YR', 'YT', 'YARD POSITION']
    all_result_cols = (core_cols + position_cols) if container_data_available else \
                      ['MOVE HOUR', 'STS', 'VESSEL BAY', 'BAY',
                       'ASSIGNED BLOCK', 'WEIGHT CLASS', 'QUANTITIES']
    col_widths_map = {
        'MOVE HOUR': 14, 'CONT LIST': 45, 'CONTAINER ID': 20,
        'ST': 10, 'POD': 10, 'STS': 10,
        'VESSEL BAY': 12,
        'BAY': 10,
        'ASSIGNED BLOCK': 16, 'WEIGHT CLASS': 14, 'QUANTITIES': 12,
        'YB': 8, 'YR': 8, 'YT': 8, 'YARD POSITION': 18
    }
    CONT_LIST_SET = {'CONT LIST'}
    CONT_ID_SET   = {'CONTAINER ID', 'ST', 'POD'}
    VESSEL_BAY_SET = {'VESSEL BAY'}
    POS_SET       = set(position_cols)
    INT_COLS      = {'YB', 'YR', 'YT'}

    def _hdr_key(cn):
        if cn in CONT_LIST_SET:   return 'hdr_pale'
        if cn in CONT_ID_SET:     return 'hdr_mid'
        if cn in VESSEL_BAY_SET:  return 'hdr_mid'
        if cn in POS_SET:         return 'hdr_light'
        return 'hdr_dark'

    _dummy = wb.create_sheet('_dummy_styles')
    def _get_sid(key):
        c = _dummy.cell(row=1, column=list(_ST.keys()).index(key) + 1)
        _apply(c, key)
        return c._style
    _sid = {k: _get_sid(k) for k in _ST}
    wb.remove(_dummy)

    cl_idx = (all_result_cols.index('CONT LIST') + 1
              if 'CONT LIST' in all_result_cols else None)

    def write_result_sheet(ws, df, sheet_title):
        n_cols = len(all_result_cols)
        n_rows = len(df)

        ws.append(all_result_cols)
        for ci, cn in enumerate(all_result_cols, 1):
            cell = ws.cell(row=1, column=ci)
            _apply(cell, _hdr_key(cn))
            ws.column_dimensions[get_column_letter(ci)].width = \
                col_widths_map.get(cn, 14)

        cont_list_map = {}
        if cl_idx and 'CONTAINER ID' in df.columns:
            for (mh, bay), grp in df.groupby(['MOVE HOUR', 'BAY']):
                ids = [str(v).strip() for v in grp['CONTAINER ID']
                       if str(v).strip() not in ('', 'nan')]
                cont_list_map[(mh, bay)] = ', '.join(ids) if ids else ''

        records = df.to_dict('records')

        data_sids = []
        shade = True; prev_gk = None
        for rec in records:
            gk = (rec.get('MOVE HOUR'), rec.get('STS'), rec.get('BAY'),
                  rec.get('ASSIGNED BLOCK'), rec.get('WEIGHT CLASS'))
            if gk != prev_gk:
                shade = not shade; prev_gk = gk
            sk = 'data_w' if shade else 'data_a'
            data_sids.append(_sid[sk])

            row_vals = []
            for cn in all_result_cols:
                if cn == 'CONT LIST':
                    row_vals.append(None); continue
                v = rec.get(cn, '')
                if cn in INT_COLS and v != '':
                    try: v = int(v)
                    except: pass
                if v == '' or (isinstance(v, float) and str(v) == 'nan'):
                    v = None
                row_vals.append(v)
            ws.append(row_vals)

        for ri, sid in enumerate(data_sids, 2):
            for ci in range(1, n_cols + 1):
                if cl_idx and ci == cl_idx:
                    continue
                ws.cell(row=ri, column=ci)._style = sid

        ws.sheet_format.defaultRowHeight = 28
        ws.sheet_format.customHeight = True

        if cl_idx:
            prev_key = None; grp_start = 2; n_ids_grp = 0
            def _flush_cl(pk, gs, ge, nids):
                text = cont_list_map.get(pk) or None
                if ge > gs:
                    ws.merge_cells(start_row=gs, start_column=cl_idx,
                                   end_row=ge,   end_column=cl_idx)
                top_cell = ws.cell(row=gs, column=cl_idx)
                top_cell.value = text
                top_cell._style = _sid['cont_list']

            for ri, rec in enumerate(records, 2):
                cur_key = (rec.get('MOVE HOUR', ''), rec.get('BAY', ''))
                nids = len([x for x in cont_list_map.get(cur_key, '').split(',')
                            if x.strip()])
                if cur_key != prev_key:
                    if prev_key is not None:
                        _flush_cl(prev_key, grp_start, ri - 1, n_ids_grp)
                    prev_key = cur_key; grp_start = ri; n_ids_grp = nids
            if prev_key is not None:
                _flush_cl(prev_key, grp_start, n_rows + 1, n_ids_grp)

        _log(f"  ✅ Sheet '{sheet_title}': {n_rows} dòng")

    if container_data_available and 'ST' in df_result_detail.columns:
        st_values = [s for s in sorted(df_result_detail['ST'].dropna().unique())
                     if str(s).strip() not in ('', 'nan')]
    else:
        st_values = ['ALL']
    if not st_values:
        st_values = ['ALL']

    for st_idx, st_val in enumerate(st_values, 1):
        sname = (f"RESULT {st_idx} ({st_val})"
                 if st_val != 'ALL' else 'RESULT')[:31]
        ws = wb.create_sheet(sname)
        df_rd = (df_result_detail if st_val == 'ALL' else
                 df_result_detail[
                     df_result_detail['ST'].astype(str).str.strip()
                     == str(st_val).strip()]
                 ).reset_index(drop=True)
        write_result_sheet(ws, df_rd, sname)

    ws_total = wb.create_sheet('RESULT TOTAL')
    write_result_sheet(ws_total,
                       df_result_detail.reset_index(drop=True),
                       'RESULT TOTAL')

    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)

    total_rows = len(df_result_detail)
    _log(f"✅ Hoàn thành. Rows={total_rows}, Total Clashes={total_clashes}")
    return excel_buffer, total_rows, total_clashes


# ============================================================
# Chạy trực tiếp từ dòng lệnh (không cần Streamlit)
# ============================================================
if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Cách dùng: python ALLOCATION.py <đường_dẫn_file_excel>")
        sys.exit(1)

    file_path = sys.argv[1]
    try:
        with open(file_path, "rb") as f:
            excel_buffer, total_rows, total_clashes = run_optimization(f)
    except Exception as e:
        print(f"Lỗi khi tối ưu: {e}")
        sys.exit(1)

    output_path = file_path.replace(".xlsx", " result.xlsx").replace(".xls", ".xls")
    with open(output_path, "wb") as out_f:
        out_f.write(excel_buffer.read())

    print(f"\n✅ Thành công! File kết quả: {output_path}")
    print(f"   Tổng số dòng phân bổ: {total_rows}")
    print(f"   Tổng số clash:        {total_clashes}")
